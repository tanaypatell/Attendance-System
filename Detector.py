import cv2
from time import sleep
from PIL import Image 
import openpyxl
from datetime import datetime
from tkinter import messagebox
itera=2

def main_app(name):
        global itera
        chin,chout=False,False
        face_cascade = cv2.CascadeClassifier('./data/haarcascade_frontalface_default.xml')
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read(f"./data/classifiers/{name}_classifier.xml")
        cap = cv2.VideoCapture(0)
        pred = 0
        while True:
            ret, frame = cap.read()
            #default_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray,1.3,5)

            for (x,y,w,h) in faces:


                roi_gray = gray[y:y+h,x:x+w]

                id,confidence = recognizer.predict(roi_gray)
                confidence = 100 - int(confidence)
                pred = 0
                if confidence > 50:
                    #if u want to print confidence level
                            #confidence = 100 - int(confidence)
                            pred += +1
                            text = name.upper()
                            font = cv2.FONT_HERSHEY_PLAIN
                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                            frame = cv2.putText(frame, text+" Recognized", (x, y-4), font, 1, (0, 255, 0), 1, cv2.LINE_AA)

                else:   
                            pred += -1
                            text = "UnknownFace, please try again"
                            font = cv2.FONT_HERSHEY_PLAIN
                            frame = cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                            frame = cv2.putText(frame, text, (x, y-4), font, 1, (0, 0,255), 1, cv2.LINE_AA)

            cv2.imshow("image", frame)


            if cv2.waitKey(20) & 0xFF == ord('q'):
                print(pred)
                if pred > 0 : 
                    dim =(124,124)
                    img = cv2.imread(f".\\data\\{name}\\{pred}{name}.jpg", cv2.IMREAD_UNCHANGED)
                    resized = cv2.resize(img, dim, interpolation = cv2.INTER_AREA)
                    cv2.imwrite(f".\\data\\{name}\\50{name}.jpg", resized)
                    Image1 = Image.open(f".\\2.png") 
                      
                    # make a copy the image so that the  
                    # original image does not get affected 
                    Image1copy = Image1.copy() 
                    Image2 = Image.open(f".\\data\\{name}\\50{name}.jpg") 
                    Image2copy = Image2.copy() 
                      
                    # paste image giving dimensions 
                    Image1copy.paste(Image2copy, (195, 114)) 
                      
                    # save the image  
                    Image1copy.save("end.png") 
                    frame = cv2.imread("end.png", 1)

                    cv2.imshow("Result",frame)


                    wb = openpyxl.load_workbook("Attendance.xlsx")

                    now = datetime.now()
                    current_time = now.strftime("%H:%M:%S")
                    while True:
                        if wb.worksheets[0].cell(row=itera,column=1).value:
                            if wb.worksheets[0].cell(row=itera,column=1).value==name and wb.worksheets[0].cell(row=itera,column=3).value==None:
                                wb.worksheets[0].cell(row=itera,column=3).value=current_time
                                chout=True
                                # messagebox.showinfo("SUCCESS", "Check-out time successfully recorded")
                                # itera+=1
                                break
                            itera+=1
                            continue
                        else:
                            wb.worksheets[0].cell(row=itera,column=1).value=name
                            wb.worksheets[0].cell(row=itera,column=2).value=current_time
                            chin=True
                            # itera=2
                            # messagebox.showinfo("SUCCESS", "Check-in time successfully recorded")
                            # itera+=1
                            break

                                                
                    wb.save("Attendance.xlsx")
                    wb.close
                    # itera+=1
                 
                    cv2.waitKey(2500)
                    if chin:
                        messagebox.showinfo("SUCCESS", "Check-in time successfully recorded")
                        chin=False
                    if chout:
                        messagebox.showinfo("SUCCESS", "Check-out time successfully recorded")
                        chout=False
                break


        cap.release()
        cv2.destroyAllWindows()
        
